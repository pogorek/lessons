# 13 WORKING WITH EXCEL SPREADSHEETS

## Opening Excel Documents with OpenPyXL

# import openpyxl
# wb = openpyxl.load_workbook('example.xlsx')
# print(type(wb))
# # <class 'openpyxl.workbook.workbook.Workbook'>

## Getting Sheets from the Workbook

import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
print(wb.sheetnames) # The workbook's sheets' names.
# ['Sheet1', 'Sheet2', 'Sheet3']
sheet = wb['Sheet3'] # Get a sheet from the workbook.
print(sheet)
# <Worksheet "Sheet3">
print(type(sheet))
# <class 'openpyxl.worksheet.worksheet.Worksheet'>
print(sheet.title) # Get the sheet's title as a string.
# 'Sheet3'
anotherSheet = wb.active # Get the active sheet.
print(anotherSheet)
# <Worksheet "Sheet1">

## Getting Cells from the Sheets